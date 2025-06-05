import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="Clasificador de Piezas", layout="wide")
st.title("🔩 Clasificación de Piezas por Longitud")

# === SUBIR ARCHIVO ===
archivo_cargado = st.file_uploader("Sube el archivo 'Datos.xlsx'", type=["xlsx"])

# === FUNCIÓN AUXILIAR PARA NOMBRES DE HOJA ===
def get_unique_sheet_name(base_name, used_names):
    base = str(base_name)[:31] if base_name else "Hoja"
    name = base
    count = 1
    while name in used_names:
        suffix = f"_{count}"
        name = (base[:31 - len(suffix)] + suffix)
        count += 1
    used_names.add(name)
    return name

if archivo_cargado:
    df = pd.read_excel(archivo_cargado)
    df = df[['OBRA', 'ID', 'Size', 'Length (m)']].dropna()
    df['Length (m)'] = df['Length (m)'].round(2)

    longitudes_estandar = [1.0, 2.0, 2.4, 3.0, 4.0, 6.0, 12.0]
    df_estandar = df[df['Length (m)'].isin(longitudes_estandar)].copy()
    df_sobrantes = df[~df['Length (m)'].isin(longitudes_estandar)].copy()
    df_mayores_12 = df_sobrantes[df_sobrantes['Length (m)'] > 12.0].copy()
    grouped_estandar = dict(tuple(df_estandar.groupby('Size')))
    grouped_sobrantes = dict(tuple(df_sobrantes[df_sobrantes['Length (m)'] <= 12.0].groupby('Size')))

    # === TABLAS PARA MOSTRAR ===
    st.subheader("📊 Resumen general de piezas estándar")
    res = df_estandar.groupby(['Size', 'Length (m)']).size().reset_index(name='Total Piezas')
    st.dataframe(res)

    st.subheader("📊 Resumen general de piezas sobrantes")
    df_resumen_size = df_sobrantes.groupby('Size').agg(
        Total=('Length (m)', 'count'),
        Menor_Igual_12=('Length (m)', lambda x: (x <= 12.0).sum()),
        Mayor_12=('Length (m)', lambda x: (x > 12.0).sum())
    ).reset_index()
    st.dataframe(df_resumen_size)

    # === CREACIÓN DE ARCHIVOS ===
    def crear_archivos():
        output_files = {}
        used_names_estandar = set()
        used_names_sobrantes = set()

        # piezas_estandar.xlsx
        wb_estandar = Workbook()
        ws_general = wb_estandar.active
        ws_general.title = "Resumen General"
        used_names_estandar.add(ws_general.title)
        ws_general.append(['Size', 'Length (m)', 'Total Piezas'])
        for _, row in res.iterrows():
            ws_general.append(row.tolist())

        ws_id = wb_estandar.create_sheet("Resumen por ID")
        used_names_estandar.add(ws_id.title)
        ws_id.append(['Size', 'ID', 'Length (m)', 'Frecuencia'])
        res_id = df_estandar.groupby(['Size', 'ID', 'Length (m)']).size().reset_index(name='Frecuencia')
        for _, row in res_id.iterrows():
            ws_id.append(row.tolist())

        for size, group in grouped_estandar.items():
            title = get_unique_sheet_name(size, used_names_estandar)
            ws = wb_estandar.create_sheet(title=title)
            ws.append(['OBRA', 'ID', 'Length (m)'])
            for _, row in group.iterrows():
                ws.append([row['OBRA'], row['ID'], row['Length (m)']])

        buffer1 = BytesIO()
        wb_estandar.save(buffer1)
        output_files["piezas_estandar.xlsx"] = buffer1

        # piezas_sobrantes.xlsx
        wb_sobrantes = Workbook()
        ws_resumen_general = wb_sobrantes.active
        ws_resumen_general.title = "Resumen size"
        used_names_sobrantes.add(ws_resumen_general.title)
        ws_resumen_general.append(["Size", "Total piezas", "Piezas <= 12m", "Piezas > 12m"])
        for _, row in df_resumen_size.iterrows():
            ws_resumen_general.append(row.tolist())

        ws_id_sob = wb_sobrantes.create_sheet("Resumen por ID")
        used_names_sobrantes.add(ws_id_sob.title)
        ws_id_sob.append(['Size', 'ID', 'Length (m)', 'Frecuencia'])
        res_id_sob = df_sobrantes.groupby(['Size', 'ID', 'Length (m)']).size().reset_index(name='Frecuencia')
        for _, row in res_id_sob.iterrows():
            ws_id_sob.append(row.tolist())

        ws_mayores = wb_sobrantes.create_sheet("Mayores a 12m")
        used_names_sobrantes.add(ws_mayores.title)
        ws_mayores.append(['OBRA', 'ID', 'Size', 'Length (m)', 'Comentario'])
        for _, row in df_mayores_12.iterrows():
            ws_mayores.append([row['OBRA'], row['ID'], row['Size'], row['Length (m)'], 'No optimizable'])

        for size, group in grouped_sobrantes.items():
            title = get_unique_sheet_name(size, used_names_sobrantes)
            ws = wb_sobrantes.create_sheet(title=title)
            ws.append(['OBRA', 'ID', 'Length (m)'])
            for _, row in group.iterrows():
                ws.append([row['OBRA'], row['ID'], row['Length (m)']])

        buffer2 = BytesIO()
        wb_sobrantes.save(buffer2)
        output_files["piezas_sobrantes.xlsx"] = buffer2

        # Datos-y-Resumen.xlsx
        res_est = df_estandar.groupby('Size').size().reset_index(name='Total pieza')
        buffer3 = BytesIO()
        with pd.ExcelWriter(buffer3, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Datos Originales", index=False)
            wb = writer.book
            title_final = get_unique_sheet_name("Resumen Estandar y Sobrante", set(wb.sheetnames))
            ws = wb.create_sheet(title=title_final)
            ws.append(["Estandar"])
            ws.append(["Size", "Total pieza"])
            for _, row in res_est.iterrows():
                ws.append(row.tolist())
            ws.append([])
            ws.append(["Sobrante"])
            ws.append(["Size", "Total piezas", "Piezas <= 12m", "Piezas > 12m"])
            for _, row in df_resumen_size.iterrows():
                ws.append(row.tolist())
        output_files["Datos-y-Resumen.xlsx"] = buffer3

        return output_files

    files = crear_archivos()

    # === BOTONES DE DESCARGA CON FECHA Y HORA ===
    st.subheader("📥 Descarga de archivos generados")
    timestamp = datetime.now().strftime("%d%m%y-%H%M_")
    for filename, buffer in files.items():
        st.download_button(
            label=f"Descargar {filename}",
            data=buffer.getvalue(),
            file_name=timestamp + filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# === PIE DE PÁGINA CON CRÉDITO ===
st.markdown(
    "<br><hr><div style='text-align: center; font-size: 0.9em;'>"
    "Desarrollado por <strong>Dra. J. Isabel Méndez</strong> para <strong>CIMESA</strong>"
    "</div>",
    unsafe_allow_html=True
)
